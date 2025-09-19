import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import logging
import os
from decouple import config  # For .env file support
import requests
from datetime import datetime
import time
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Color
import yfinance as yf
from dateutil.relativedelta import relativedelta

# Setup logging
USER_DATA_DIR = os.path.expanduser("~/PortfolioTracker")
if not os.path.exists(USER_DATA_DIR):
    os.makedirs(USER_DATA_DIR)
logging.basicConfig(
    filename=os.path.join(USER_DATA_DIR, 'portfolio.log'),
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger()

class PortfolioTrackerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Portfolio Tracker")
        self.root.geometry("1400x900")
        self.fmp_api_key = config('FMP_API_KEY')  # Load from .env
        self.fred_api_key = config('FRED_API_KEY')  # Load from .env
        self.dark_mode = False

        # Styling
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.apply_light_theme()

        # Create main frame
        self.main_frame = ttk.Frame(self.root)
        self.main_frame.pack(expand=True, fill="both", padx=5, pady=5)
        self.main_frame.grid_columnconfigure(0, weight=1)
        self.main_frame.grid_rowconfigure(0, weight=0)
        self.main_frame.grid_rowconfigure(1, weight=1)
        self.main_frame.grid_rowconfigure(2, weight=0)
        self.main_frame.grid_rowconfigure(3, weight=1)

        # Summary label
        self.summary_label = ttk.Label(self.main_frame, text="Portfolio Summary: Loading...", style="Summary.TLabel")
        self.summary_label.grid(row=0, column=0, columnspan=2, sticky="ew", pady=5)

        # Create treeview
        self.tree = ttk.Treeview(
            self.main_frame,
            columns=("Symbol", "Name", "Purchase Date", "Purchase Price", "Shares", "Current Price", "Value", "Gain/Loss", "Intrinsic", "Margin", "Alert Threshold"),
            show="headings"
        )
        self.tree.heading("Symbol", text="Symbol")
        self.tree.heading("Name", text="Company Name")
        self.tree.heading("Purchase Date", text="Purchase Date")
        self.tree.heading("Purchase Price", text="Purchase Price ($)")
        self.tree.heading("Shares", text="Shares Owned")
        self.tree.heading("Current Price", text="Current Price ($)")
        self.tree.heading("Value", text="Total Value ($)")
        self.tree.heading("Gain/Loss", text="Gain/Loss ($)")
        self.tree.heading("Intrinsic", text="Intrinsic Value ($)")
        self.tree.heading("Margin", text="Margin of Safety (%)")
        self.tree.heading("Alert Threshold", text="Alert Threshold ($)")
        self.tree.column("Symbol", width=80, anchor="center")
        self.tree.column("Name", width=200, anchor="center")  # Centered
        self.tree.column("Purchase Date", width=100, anchor="center")
        self.tree.column("Purchase Price", width=120, anchor="center")
        self.tree.column("Shares", width=80, anchor="center")
        self.tree.column("Current Price", width=120, anchor="center")
        self.tree.column("Value", width=120, anchor="center")
        self.tree.column("Gain/Loss", width=100, anchor="center")
        self.tree.column("Intrinsic", width=120, anchor="center")
        self.tree.column("Margin", width=120, anchor="center")
        self.tree.column("Alert Threshold", width=120, anchor="center")
        self.tree.grid(row=1, column=0, sticky="nsew")

        # Scrollbar
        scrollbar = ttk.Scrollbar(self.main_frame, orient="vertical", command=self.tree.yview)
        scrollbar.grid(row=1, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar.set)

        # Add stock entry
        self.entry_frame = ttk.Frame(self.main_frame)
        self.entry_frame.grid(row=2, column=0, sticky="ew")
        ttk.Label(self.entry_frame, text="Symbol:").pack(side="left")
        self.symbol_entry = ttk.Entry(self.entry_frame, width=10)
        self.symbol_entry.pack(side="left", padx=5)
        ttk.Label(self.entry_frame, text="Shares:").pack(side="left")
        self.shares_entry = ttk.Entry(self.entry_frame, width=10)
        self.shares_entry.pack(side="left", padx=5)
        ttk.Label(self.entry_frame, text="Purchase Date (YYYY-MM-DD):").pack(side="left")
        self.date_entry = ttk.Entry(self.entry_frame, width=12)
        self.date_entry.pack(side="left", padx=5)
        ttk.Label(self.entry_frame, text="Purchase Price ($):").pack(side="left")
        self.purchase_price_entry = ttk.Entry(self.entry_frame, width=10)
        self.purchase_price_entry.pack(side="left", padx=5)
        ttk.Label(self.entry_frame, text="Alert Threshold ($):").pack(side="left")
        self.alert_threshold_entry = ttk.Entry(self.entry_frame, width=10)
        self.alert_threshold_entry.pack(side="left", padx=5)
        ttk.Button(self.entry_frame, text="Add Stock", command=self.add_stock).pack(side="left", padx=5)
        ttk.Button(self.entry_frame, text="Refresh Prices", command=self.refresh_prices).pack(side="left", padx=5)
        ttk.Button(self.entry_frame, text="Show Chart", command=self.show_chart).pack(side="left", padx=5)
        ttk.Button(self.entry_frame, text="Toggle Dark Mode", command=self.toggle_theme).pack(side="left", padx=5)
        ttk.Button(self.entry_frame, text="Export to Excel", command=self.export_to_excel).pack(side="left", padx=5)
        ttk.Button(self.entry_frame, text="Clear Portfolio", command=self.clear_portfolio).pack(side="left", padx=5)

        # Chart frame
        self.chart_frame = ttk.Frame(self.main_frame)
        self.chart_frame.grid(row=3, column=0, columnspan=2, sticky="nsew")
        self.chart_canvas = None

        # Initialize database and load portfolio
        self.init_db()
        self.load_portfolio()

    def apply_light_theme(self):
        """Apply light theme styling."""
        self.style.configure("TButton", font=("Helvetica", 10), padding=5, background="#4CAF50", foreground="white")
        self.style.configure("TLabel", font=("Helvetica", 10), background="#f0f0f0")
        self.style.configure("Summary.TLabel", font=("Helvetica", 12, "bold"), background="#e0e0e0")
        self.root.configure(bg="#f0f0f0")

    def apply_dark_theme(self):
        """Apply dark theme styling."""
        self.style.configure("TButton", font=("Helvetica", 10), padding=5, background="#2196F3", foreground="white")
        self.style.configure("TLabel", font=("Helvetica", 10), background="#333333", foreground="white")
        self.style.configure("Summary.TLabel", font=("Helvetica", 12, "bold"), background="#444444", foreground="white")
        self.root.configure(bg="#333333")

    def toggle_theme(self):
        """Toggle between light and dark themes."""
        self.dark_mode = not self.dark_mode
        if self.dark_mode:
            self.apply_dark_theme()
        else:
            self.apply_light_theme()
        if self.chart_canvas:
            self.show_chart()
        logger.info(f"Toggled to {'dark' if self.dark_mode else 'light'} mode")

    def init_db(self):
        """Initialize SQLite database for portfolio and history."""
        conn = sqlite3.connect(os.path.join(USER_DATA_DIR, "portfolio.db"))
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS portfolio (
                symbol TEXT PRIMARY KEY,
                company_name TEXT,
                purchase_date TEXT,
                purchase_price REAL,
                shares INTEGER,
                price REAL,
                eps_ttm REAL,
                eps_cagr REAL,
                intrinsic_value REAL,
                alert_threshold REAL,
                last_updated TEXT
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS portfolio_history (
                date TEXT,
                total_value REAL
            )
        """)
        conn.commit()
        conn.close()

    def load_portfolio(self):
        """Load portfolio from database into treeview and update summary."""
        conn = sqlite3.connect(os.path.join(USER_DATA_DIR, "portfolio.db"))
        cursor = conn.cursor()
        cursor.execute("SELECT symbol, company_name, purchase_date, purchase_price, shares, price, intrinsic_value, alert_threshold FROM portfolio")
        rows = cursor.fetchall()
        conn.close()

        # Clear existing treeview entries
        for item in self.tree.get_children():
            self.tree.delete(item)

        total_value = 0
        total_gain_loss = 0
        total_margin = 0
        valid_rows = 0
        for symbol, name, purchase_date, purchase_price, shares, price, intrinsic_value, alert_threshold in rows:
            if not all([price, shares]) or purchase_price is None:  # Allow $0 purchase_price for splits
                logger.warning(f"Skipping {symbol} due to missing data: price={price}, purchase_price={purchase_price}, shares={shares}")
                continue
            value = price * shares
            gain_loss = (price - purchase_price) * shares
            margin = ((intrinsic_value - price) / intrinsic_value * 100) if intrinsic_value and price and intrinsic_value > 0 else 0
            self.tree.insert("", "end", values=(
                symbol, name or "N/A", purchase_date or "N/A", f"${purchase_price:.2f}" if purchase_price is not None else "$0.00", shares, f"${price:.2f}", f"${value:.2f}",
                f"${gain_loss:.2f}" if gain_loss > 0 else f"(${abs(gain_loss):.2f})", f"${intrinsic_value:.2f}" if intrinsic_value else "N/A",
                f"{margin:.1f}%" if margin else "N/A", f"${alert_threshold:.2f}" if alert_threshold else "N/A"
            ))
            total_value += value
            total_gain_loss += gain_loss
            total_margin += margin  # Include 0 margins for average
            valid_rows += 1
            if alert_threshold and price and abs(price - alert_threshold) <= 0.05 * alert_threshold:
                messagebox.showinfo("Price Alert", f"{symbol} price (${price:.2f}) is near alert threshold (${alert_threshold:.2f})")
        avg_margin = total_margin / valid_rows if valid_rows > 0 else 0
        self.summary_label.config(text=f"Portfolio Summary: {valid_rows} stocks, Total Value: ${total_value:.2f}, Total Gain/Loss: ${total_gain_loss:.2f}, Avg Margin of Safety: {avg_margin:.1f}%")
        logger.info(f"Loaded portfolio, Total Value: ${total_value:.2f}, Total Gain/Loss: ${total_gain_loss:.2f}, Valid Stocks: {valid_rows}")

        if valid_rows > 0:
            self.save_portfolio_value(total_value)

    def add_stock(self):
        """Add a stock to the portfolio."""
        symbol = self.symbol_entry.get().strip().upper()
        shares = self.shares_entry.get().strip()
        purchase_date = self.date_entry.get().strip()
        purchase_price = self.purchase_price_entry.get().strip()
        alert_threshold = self.alert_threshold_entry.get().strip()
        if not symbol or not shares.isdigit() or not purchase_date or not purchase_price.replace('.', '').isdigit():
            messagebox.showerror("Error", "Enter a valid symbol, number of shares, purchase date (YYYY-MM-DD), and purchase price")
            logger.error("Invalid input: All fields required")
            return

        shares = int(shares)
        purchase_price = float(purchase_price)
        alert_threshold = float(alert_threshold) if alert_threshold.replace('.', '').isdigit() else None
        if not self.is_valid_date(purchase_date):
            messagebox.showerror("Error", "Purchase date must be in YYYY-MM-DD format")
            logger.error("Invalid purchase date format")
            return

        price, name, eps_ttm, eps_cagr = self.fetch_stock_data(symbol)
        if price is None or name is None:
            messagebox.showerror("Error", f"Failed to fetch data for {symbol}: Check ticker or network connection")
            logger.error(f"Fetch failed for {symbol}: price={price}, name={name}")
            return

        intrinsic_value = self.calculate_graham_value(eps_ttm, eps_cagr) if eps_ttm and eps_cagr else None

        # Save to database
        conn = sqlite3.connect(os.path.join(USER_DATA_DIR, "portfolio.db"))
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM portfolio WHERE symbol = ?", (symbol,))
        if cursor.fetchone()[0] > 0:
            logger.warning(f"Stock {symbol} already exists, updating instead")
            cursor.execute("""
                UPDATE portfolio SET company_name = ?, purchase_date = ?, purchase_price = ?, shares = ?, price = ?, eps_ttm = ?, eps_cagr = ?, intrinsic_value = ?, alert_threshold = ?, last_updated = ?
                WHERE symbol = ?
            """, (name, purchase_date, purchase_price, shares, price, eps_ttm, eps_cagr, intrinsic_value, alert_threshold, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), symbol))
        else:
            cursor.execute("""
                INSERT INTO portfolio (symbol, company_name, purchase_date, purchase_price, shares, price, eps_ttm, eps_cagr, intrinsic_value, alert_threshold, last_updated)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (symbol, name, purchase_date, purchase_price, shares, price, eps_ttm, eps_cagr, intrinsic_value, alert_threshold, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        conn.commit()
        conn.close()

        # Recalculate total_value for the new stock
        value = price * shares
        conn = sqlite3.connect(os.path.join(USER_DATA_DIR, "portfolio.db"))
        cursor = conn.cursor()
        cursor.execute("SELECT SUM(price * shares) FROM portfolio")
        total_value = cursor.fetchone()[0] or 0
        conn.close()

        # Update treeview and reload portfolio
        self.load_portfolio()  # Clears and reloads treeview
        intrinsic_str = f"{intrinsic_value:.2f}" if intrinsic_value is not None else "N/A"
        logger.info(f"Added/Updated {symbol} with {shares} shares at purchase ${purchase_price:.2f}, current ${price:.2f}, intrinsic {intrinsic_str}")

        self.symbol_entry.delete(0, tk.END)
        self.shares_entry.delete(0, tk.END)
        self.date_entry.delete(0, tk.END)
        self.purchase_price_entry.delete(0, tk.END)
        self.alert_threshold_entry.delete(0, tk.END)

        # Update portfolio history
        self.save_portfolio_value(total_value)

    def is_valid_date(self, date_str):
        """Validate date format YYYY-MM-DD."""
        try:
            datetime.strptime(date_str, "%Y-%m-%d")
            return True
        except ValueError:
            return False

    def refresh_prices(self):
        """Refresh stock prices in the portfolio."""
        conn = sqlite3.connect(os.path.join(USER_DATA_DIR, "portfolio.db"))
        cursor = conn.cursor()
        cursor.execute("SELECT symbol, shares, purchase_price, alert_threshold FROM portfolio")
        rows = cursor.fetchall()

        # Clear treeview
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Update prices with rate limiting
        total_value = 0
        total_gain_loss = 0
        total_margin = 0
        count = 0
        for i, (symbol, shares, purchase_price, alert_threshold) in enumerate(rows):
            price, name, eps_ttm, eps_cagr = self.fetch_stock_data(symbol)
            if i % 5 == 0:  # Simple rate limit (5 calls per batch)
                time.sleep(1)
            if price is None or name is None:
                logger.error(f"Failed to refresh data for {symbol}: price={price}, name={name}")
                continue
            intrinsic_value = self.calculate_graham_value(eps_ttm, eps_cagr) if eps_ttm and eps_cagr else None
            cursor.execute("""
                UPDATE portfolio SET price = ?, company_name = ?, eps_ttm = ?, eps_cagr = ?, intrinsic_value = ?, last_updated = ?
                WHERE symbol = ?
            """, (price, name, eps_ttm, eps_cagr, intrinsic_value, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), symbol))
            value = price * shares
            gain_loss = (price - purchase_price) * shares if purchase_price and price and shares else 0
            margin = ((intrinsic_value - price) / intrinsic_value * 100) if intrinsic_value and price and intrinsic_value > 0 else 0
            self.tree.insert("", "end", values=(
                symbol, name or "N/A", "N/A" if not purchase_price else "N/A", f"${purchase_price:.2f}" if purchase_price is not None else "$0.00", shares, f"${price:.2f}", f"${value:.2f}",
                f"${gain_loss:.2f}" if gain_loss > 0 else f"(${abs(gain_loss):.2f})", f"${intrinsic_value:.2f}" if intrinsic_value else "N/A",
                f"{margin:.1f}%" if margin else "N/A", f"${alert_threshold:.2f}" if alert_threshold else "N/A"
            ))
            total_value += value
            total_gain_loss += gain_loss
            if margin is not None:  # Include all margins for average
                total_margin += margin
                count += 1
            if alert_threshold and price and abs(price - alert_threshold) <= 0.05 * alert_threshold:
                messagebox.showinfo("Price Alert", f"{symbol} price (${price:.2f}) is near alert threshold (${alert_threshold:.2f})")
        conn.commit()
        conn.close()
        avg_margin = total_margin / count if count > 0 else 0
        self.summary_label.config(text=f"Portfolio Summary: {count} stocks, Total Value: ${total_value:.2f}, Total Gain/Loss: ${total_gain_loss:.2f}, Avg Margin of Safety: {avg_margin:.1f}%")
        logger.info(f"Refreshed {len(rows)} stocks, Total Value: ${total_value:.2f}, Total Gain/Loss: ${total_gain_loss:.2f}, Valid Stocks: {count}")

        # Update portfolio history
        self.save_portfolio_value(total_value)

    def save_portfolio_value(self, total_value):
        """Save total portfolio value to history."""
        conn = sqlite3.connect(os.path.join(USER_DATA_DIR, "portfolio.db"))
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO portfolio_history (date, total_value)
            VALUES (?, ?)
        """, (datetime.now().strftime("%Y-%m-%d"), total_value))
        conn.commit()
        conn.close()
        logger.info(f"Saved portfolio value: ${total_value:.2f}")

    def show_chart(self):
        """Display a chart of portfolio value vs benchmarks."""
        conn = sqlite3.connect(os.path.join(USER_DATA_DIR, "portfolio.db"))
        cursor = conn.cursor()
        cursor.execute("SELECT date, total_value FROM portfolio_history ORDER BY date")
        portfolio_rows = cursor.fetchall()
        conn.close()

        if not portfolio_rows:
            messagebox.showinfo("No Data", "No portfolio history available")
            return

        dates, portfolio_values = zip(*portfolio_rows)
        start_date = datetime.strptime(dates[0], "%Y-%m-%d") - relativedelta(months=1)
        end_date = datetime.strptime(dates[-1], "%Y-%m-%d") + relativedelta(days=1)

        # Fetch benchmark data with yfinance
        benchmarks = {
            "S&P 500": "^GSPC",
            "NYSE Composite": "^NYA",
            "NASDAQ Composite": "^IXIC"
        }
        benchmark_data = {}
        for name, ticker in benchmarks.items():
            try:
                data = yf.download(ticker, start=start_date, end=end_date)['Close']
                benchmark_data[name] = (data.index.strftime("%Y-%m-%d").tolist(), data.values)
            except Exception as e:
                logger.error(f"Error fetching {name} data: {str(e)}")
                benchmark_data[name] = ([], [])

        # Normalize portfolio values to start at 100 for comparison
        portfolio_base = portfolio_values[0] if portfolio_values else 100
        normalized_portfolio = [100 * v / portfolio_base for v in portfolio_values]

        # Create chart
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.plot(dates, normalized_portfolio, marker='o', label='Portfolio', color='#1f77b4' if not self.dark_mode else '#64B5F6')
        colors = ['#ff7f0e', '#2ca02c', '#d62728'] if not self.dark_mode else ['#ffaa00', '#66cc66', '#ff6666']
        for (name, (b_dates, b_values)), color in zip(benchmark_data.items(), colors):
            if b_values.size > 0:
                base_value = b_values[0] if b_values.size > 0 else 100
                normalized_bench = [100 * v / base_value for v in b_values]
                ax.plot(b_dates, normalized_bench, marker='s', label=name, color=color)
        ax.set_title("Portfolio vs Benchmarks (Normalized)", fontsize=14, color='black' if not self.dark_mode else 'white')
        ax.set_xlabel("Date", fontsize=12, color='black' if not self.dark_mode else 'white')
        ax.set_ylabel("Normalized Value (Base 100)", fontsize=12, color='black' if not self.dark_mode else 'white')
        ax.legend()
        ax.tick_params(axis='x', rotation=45, colors='black' if not self.dark_mode else 'white')
        ax.tick_params(axis='y', colors='black' if not self.dark_mode else 'white')
        ax.grid(True, linestyle='--', alpha=0.7, color='gray' if not self.dark_mode else 'lightgray')
        ax.set_facecolor('#f0f0f0' if not self.dark_mode else '#333333')
        fig.set_facecolor('#f0f0f0' if not self.dark_mode else '#333333')
        plt.tight_layout()

        # Embed in Tkinter
        if self.chart_canvas:
            self.chart_canvas.get_tk_widget().destroy()
        self.chart_canvas = FigureCanvasTkAgg(fig, master=self.chart_frame)
        self.chart_canvas.draw()
        self.chart_canvas.get_tk_widget().pack(fill="both", expand=True)
        self.chart_canvas.get_tk_widget().configure(bg='#f0f0f0' if not self.dark_mode else '#333333')
        plt.close(fig)
        logger.info("Displayed portfolio vs benchmarks chart")

    def export_to_excel(self):
        """Export portfolio data to an Excel file."""
        conn = sqlite3.connect(os.path.join(USER_DATA_DIR, "portfolio.db"))
        cursor = conn.cursor()
        cursor.execute("SELECT symbol, company_name, purchase_date, purchase_price, shares, price, intrinsic_value, alert_threshold FROM portfolio")
        rows = cursor.fetchall()
        cursor.execute("SELECT date, total_value FROM portfolio_history ORDER BY date")
        history_rows = cursor.fetchall()
        conn.close()

        if not rows:
            messagebox.showinfo("No Data", "No portfolio data to export")
            return

        # Create workbook
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Portfolio Summary"

        # Headers
        headers = ["Symbol", "Company Name", "Purchase Date", "Purchase Price ($)", "Shares", "Current Price ($)", "Total Value ($)", "Gain/Loss ($)", "Intrinsic Value ($)", "Margin of Safety (%)", "Alert Threshold ($)"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Data
        green_fill = PatternFill(start_color=Color(rgb="90EE90"), end_color=Color(rgb="90EE90"), fill_type="solid")
        red_fill = PatternFill(start_color=Color(rgb="FFB6C1"), end_color=Color(rgb="FFB6C1"), fill_type="solid")
        for row_idx, (symbol, name, purchase_date, purchase_price, shares, price, intrinsic_value, alert_threshold) in enumerate(rows, 2):
            value = price * shares if price and shares else 0
            gain_loss = (price - purchase_price) * shares if purchase_price and price and shares else 0
            margin = ((intrinsic_value - price) / intrinsic_value * 100) if intrinsic_value and price and intrinsic_value > 0 else 0
            sheet.cell(row=row_idx, column=1, value=symbol)
            sheet.cell(row=row_idx, column=2, value=name or "N/A")
            sheet.cell(row=row_idx, column=3, value=purchase_date or "N/A")
            sheet.cell(row=row_idx, column=4, value=purchase_price).number_format = "$#,##0.00"
            sheet.cell(row=row_idx, column=5, value=shares)
            sheet.cell(row=row_idx, column=6, value=price).number_format = "$#,##0.00"
            sheet.cell(row=row_idx, column=7, value=value).number_format = "$#,##0.00"
            gain_loss_cell = sheet.cell(row=row_idx, column=8, value=gain_loss)
            gain_loss_cell.number_format = "$#,##0.00"
            if gain_loss > 0:
                gain_loss_cell.fill = green_fill
            elif gain_loss < 0:
                gain_loss_cell.fill = red_fill
            sheet.cell(row=row_idx, column=9, value=intrinsic_value if intrinsic_value else "N/A").number_format = "$#,##0.00"
            sheet.cell(row=row_idx, column=10, value=margin if margin else "N/A").number_format = "0.0%"
            sheet.cell(row=row_idx, column=11, value=alert_threshold if alert_threshold else "N/A").number_format = "$#,##0.00"

        # Add history sheet
        history_sheet = wb.create_sheet("Portfolio History")
        history_sheet.append(["Date", "Total Value ($)"])
        for row_idx, (date, value) in enumerate(history_rows, 2):
            history_sheet.cell(row_idx, 1, value=date)
            history_sheet.cell(row_idx, 2, value=value).number_format = "$#,##0.00"

        # Add chart to history sheet
        chart = LineChart()
        chart.title = "Portfolio Value Over Time"
        chart.x_axis.title = "Date"
        chart.y_axis.title = "Value ($)"
        data = Reference(history_sheet, min_col=2, min_row=1, max_row=len(history_rows) + 1)
        dates = Reference(history_sheet, min_col=1, min_row=2, max_row=len(history_rows) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(dates)
        history_sheet.add_chart(chart, "D2")

        # Save file
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="Portfolio_Export.xlsx"
        )
        if file_path:
            wb.save(file_path)
            messagebox.showinfo("Success", f"Portfolio exported to {file_path}")
            logger.info(f"Exported portfolio to {file_path}")

    def fetch_stock_data(self, symbol):
        """Fetch stock price and name from yfinance, EPS TTM and CAGR from FMP if available."""
        try:
            # Primary source: yfinance for price and name
            stock = yf.Ticker(symbol)
            hist = stock.history(period="1mo")  # Use 1 month to ensure data
            if hist.empty:
                logger.warning(f"No 1-month data for {symbol}, trying 1-week period")
                hist = stock.history(period="1wk")
                if hist.empty:
                    logger.warning(f"No 1-week data for {symbol}, trying 1-day period")
                    hist = stock.history(period="1d")
                    if hist.empty:
                        raise ValueError(f"No data available for {symbol}")
            price = float(hist["Close"].iloc[-1])
            name = stock.info.get("longName", symbol)
            logger.debug(f"yfinance data for {symbol}: price={price}, name={name}")

            # Optional: Fetch EPS TTM and historical EPS from FMP
            url = f"https://financialmodelingprep.com/api/v3/quote/{symbol}?apikey={self.fmp_api_key}"
            response = requests.get(url)
            response.raise_for_status()
            quote_data = response.json()
            if quote_data and "eps" in quote_data[0]:
                eps_ttm = float(quote_data[0]["eps"])
            else:
                eps_ttm = stock.info.get("trailingEps", None)  # Fallback to yfinance
                logger.warning(f"No EPS TTM from FMP for {symbol}, using yfinance: {eps_ttm}")

            url = f"https://financialmodelingprep.com/api/v3/income-statement/{symbol}?apikey={self.fmp_api_key}&limit=5"
            response = requests.get(url)
            response.raise_for_status()
            income_data = response.json()
            annual_eps = [float(entry["eps"]) for entry in income_data if "eps" in entry][:5]  # Last 5 years
            if not annual_eps:
                logger.warning(f"No historical EPS from FMP for {symbol}, using yfinance fallback")
                annual_eps = [stock.info.get("trailingEps", 0)] * 5  # Fallback to current EPS
            eps_cagr = self.calculate_cagr(annual_eps[0], annual_eps[-1], len(annual_eps) - 1) if len(annual_eps) >= 2 else 0
            logger.debug(f"FMP data for {symbol}: eps_ttm={eps_ttm}, eps_cagr={eps_cagr}")

            return price, name, eps_ttm, eps_cagr
        except Exception as e:
            logger.error(f"Error fetching data for {symbol}: {str(e)}")
            return None, None, None, None

    def calculate_cagr(self, start_value, end_value, periods):
        """Calculate Compound Annual Growth Rate."""
        if not isinstance(start_value, (int, float)) or not isinstance(end_value, (int, float)):
            logger.error(f"Invalid CAGR inputs: start_value={start_value}, end_value={end_value}")
            return 0
        if start_value <= 0 or end_value <= 0 or periods <= 0:
            return 0
        try:
            return ((end_value / start_value) ** (1 / periods) - 1)
        except ZeroDivisionError:
            return 0

    def calculate_graham_value(self, eps_ttm, eps_cagr):
        """Calculate Graham intrinsic value using EPS and EPS CAGR."""
        if not eps_ttm or eps_ttm <= 0 or not eps_cagr or not isinstance(eps_cagr, (int, float)):
            logger.warning(f"Invalid Graham inputs: eps_ttm={eps_ttm}, eps_cagr={eps_cagr}")
            return None
        aaa_yield = self.get_aaa_yield()
        if aaa_yield <= 0:
            return None
        g = float(eps_cagr) * 100  # Convert to percentage, ensure float
        earnings_multiplier = min(8.5 + 2 * g, 20)  # Cap at 20
        normalization_factor = 4.4
        value = (eps_ttm * earnings_multiplier * normalization_factor) / (100 * aaa_yield)
        return value

    def get_aaa_yield(self, default_yield=0.045):
        """Fetch Moody's AAA Corporate Bond Yield from FRED."""
        try:
            url = f"https://api.stlouisfed.org/fred/series/observations?series_id=AAA&api_key={self.fred_api_key}&file_type=json&limit=1&sort_order=desc"
            response = requests.get(url)
            response.raise_for_status()
            data = response.json()
            if 'observations' in data and data['observations']:
                return float(data['observations'][0]['value']) / 100
            return default_yield
        except Exception as e:
            logger.error(f"Error fetching AAA yield: {str(e)}")
            return default_yield

    def clear_portfolio(self):
        """Clear all portfolio data from the database and treeview."""
        conn = sqlite3.connect(os.path.join(USER_DATA_DIR, "portfolio.db"))
        cursor = conn.cursor()
        cursor.execute("DELETE FROM portfolio")
        cursor.execute("DELETE FROM portfolio_history")
        conn.commit()
        conn.close()
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.summary_label.config(text="Portfolio Summary: 0 stocks, Total Value: $0.00, Total Gain/Loss: $0.00, Avg Margin of Safety: 0.0%")
        logger.info("Cleared portfolio and history")

if __name__ == "__main__":
    root = tk.Tk()
    app = PortfolioTrackerApp(root)
    root.mainloop()